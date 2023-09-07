import openpyxl
import sys

def parse_input_workbook_sheet(sheet):
    data = {}
    for row in sheet.iter_rows(min_row=4):
        if row[0].value:
            new_data = {}
            data[row[0].value] = new_data
            for cell in row[1:]:
                # Get one cell above with the date
                date = sheet.cell(cell.row -1, cell.col_idx).value
                # Get two cells above with the metric name
                metric = sheet.cell(cell.row -2, cell.col_idx).value
                value = cell.value
                if date not in new_data.keys():
                    new_data[date] = {}
                new_data[date][metric] = value
    return data


def write_output(data):
    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "output"
    # Write headers
    output_sheet.append(["Day of Month", "Date", "Site ID", "Page Views", "Unique Visitors", "Total Time Spent", "Visits", "Average Time Spent on Site"])
    for site, info in data.items():
        for date, metrics in info.items():
            output_sheet.append([date.day, date.date(), site, metrics.get("Page Views", ""), metrics.get("Unique Visitors", ""),
                                 metrics.get("Total Time Spent", ""), metrics.get("Visits", ""), metrics.get("Average Time Spent on Site", "")])
    return output_wb

if __name__ == "__main__":
    path = sys.argv[1]
    book = openpyxl.load_workbook(path)
    sheet = book["input_refresh_template"]
    parsed_input = parse_input_workbook_sheet(sheet)
    output_wb = write_output(parsed_input)
    output_wb.save("output.xlsx")
